# backend/app/services/excel_agent_service.py
# GÜNCELLENMİŞ VERSİYON: Dinamik Header Tespiti + Duplicate Sütun Düzeltme + Akıllı Analiz

import io
import csv
import traceback
from typing import Dict, Any, Optional, List
from app.services.llm_providers import get_llm_for_model
import re


# Pandas ve openpyxl import'ları
try:
    import pandas as pd
    import numpy as np
    PANDAS_AVAILABLE = True
except ImportError:
    print("UYARI: 'pandas' kütüphanesi yüklü değil. Excel analizi sınırlı olacak.")
    PANDAS_AVAILABLE = False
    pd = None

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("UYARI: 'openpyxl' kütüphanesi yüklü değil. Excel okuma çalışmayacak.")
    OPENPYXL_AVAILABLE = False
    load_workbook = None

# Tabulate kontrolü (Markdown tabloları için)
try:
    import tabulate
    TABULATE_AVAILABLE = True
except ImportError:
    TABULATE_AVAILABLE = False

# --- GÜNCELLENMİŞ PROMPTLAR ---

# --- GÜNCELLENMİŞ PROMPT (LİMİTSİZ) ---

EXCEL_AGENT_PROMPT = """Sen uzman bir Veri Analistisin. Sana bir Excel/CSV dosyasının detaylı analizi, istatistikleri ve veri önizlemesi verildi.

GÖREVİN:
Kullanıcının sorusunu, AŞAĞIDAKİ VERİLERİ kullanarak yanıtlamak.

ÖNEMLİ TARİH KURALLARI:
- "Son 3 ay" ifadesi kullanıldığında, Excel dosyasındaki AYLARI ANALİZ ET ve EN SON 3 AYI bul.
- Aylar genellikle şu sırada olur: Ocak, Şubat, Mart, Nisan, Mayıs, Haziran, Temmuz, Ağustos, Eylül, Ekim, Kasım, Aralık
- İngilizce aylar: January, February, March, April, May, June, July, August, September, October, November, December
- Kısa formlar: Jan, Feb, Mar, Apr, May, Jun, Jul, Aug, Sep, Oct, Nov, Dec
- ÖRNEK: Eğer Excel'de Ağustos, Eylül, Ekim, Kasım, Aralık varsa ve "son 3 ay" denirse, EKİM, KASIM, ARALIK'ı kullan (en son 3 ay).
- ÖRNEK: Eğer Excel'de Ocak, Şubat, Mart, Nisan varsa ve "son 3 ay" denirse, ŞUBAT, MART, NİSAN'ı kullan (en son 3 ay).
- ASLA ilk 3 ayı (Ocak, Şubat, Mart) kullanma, EN SON 3 AYI kullan.

KULLANILACAK KAYNAKLAR:
1. **"OTOMATİK GRUPLAMA ANALİZİ" (EN ÖNEMLİ):** Hazır hesaplanmış özetler buradadır.
2. **"VERİ ÖNİZLEME":** Tabloyu satır satır incelemek için burayı kullan.

CEVAP FORMATI KURALLARI (ÇOK ÖNEMLİ):
- **ASLA soruyu tekrarlama.** Doğrudan cevaba geç.
- **ASLA gereksiz açıklama yapma.** "Kullanıcının sorusu:", "Verilen Excel dosyasının..." gibi ifadeler kullanma.
- **ASLA "BÜTÇE ANALİZİ BİLGİLERİ", "VERİ ÖNİZLEME" gibi başlıklar ekleme.** Sadece tabloyu göster.
- Cevaplarını sunarken MUTLAKA **Markdown Tablosu** kullan.
- **ASLA** sana verilen "VERİ ÖNİZLEME" tablosunun tamamını kopyalayıp yapıştırma.
- Sadece sorulan soruya (örneğin "Bütçeye uygun olmayanlar") uyan satırları filtreleyerek yeni bir tablo oluştur.
- **TÜM BÖLÜMLERİ DAHIL ET:** Dosyada "Operation", "Admin" veya başka bölümler varsa, **HEPSİNİ** tabloya dahil et. Sadece bir bölümü gösterme.
- Tabloda gereksiz tekrar eden satırlar oluşturma. Her kalemden sadece bir tane olsun.
- **LİMİT YOKTUR:** Şartları sağlayan 10, 20 veya 50 kalem varsa **HEPSİNİ TABLOYA EKLE.**
- **ASLA** "ve diğerleri..." diyerek listeyi yarıda kesme. Tabloyu tam ve eksiksiz ver.
- **CEVAP FORMATI:** Sadece tabloyu göster. Başlık, açıklama veya ön bilgi ekleme.

**TABLO FORMATI KURALLARI (KRİTİK - MUTLAKA UYGULA):**
- Tabloda "nan", "NaN", "None", "null" gibi teknik değerler ASLA gösterilmemeli. Boş hücreler için "-" kullan.
- Sütun başlıkları anlamlı ve açıklayıcı olmalı. "Column_2", "nan" gibi teknik isimler kullanma.
- **BÜTÇE ANALİZİ TABLOSU FORMATI:**
  * İlk sütun: "Bölüm" (varsa) - "Operation" veya "Admin" gibi bölüm bilgisi
  * İkinci sütun: "Kalem" veya "Kategori" - Harcama kalemlerinin isimleri (örn: "IT", "3rd Party Service", "Repair & Maintenance")
  * Orta sütunlar: Aylık değerler veya index'ler (örn: "Ocak", "Şubat", "Mart" veya "Ocak Index", "Şubat Index")
    - Eğer index/yüzde ise "%" işareti kullan (örn: %154, %195)
    - Eğer tutar ise sayısal değer göster (örn: 1,234.50)
  * Son sütun: "Durum Analizi" veya "Açıklama" - Her kalem için kısa bir analiz metni
    - Örnek: "Son 3 ayda bütçeyi neredeyse ikiye katlamış."
    - Örnek: "Giderek artan bir bütçe aşımı var."
    - Örnek: "Sürekli bütçe üstünde, Ekim ayında artış göstermiş."
  * **ÖNEMLİ:** Dosyada "Operation" ve "Admin" gibi bölümler varsa, tabloda "Bölüm" sütunu ekle ve her satırın hangi bölüme ait olduğunu göster.
- Sayısal değerler düzgün formatlanmalı:
  * Yüzdeler için: %154, %195.5 gibi
  * Tutarlar için: 1,234.50 veya 1234.50 gibi
  * Ondalık sayılar için maksimum 2 basamak göster
- Tablo düzeni temiz ve okunabilir olmalı:
  * Sütunlar hizalı
  * Başlıklar net ve anlaşılır
  * Her satır bir kalemi temsil etmeli
  * Tekrarlayan satırlar olmamalı

**ÖRNEK TABLO FORMATI (Bölüm bilgisi varsa):**
| Bölüm | Kalem | Ocak Index | Şubat Index | Mart Index | Durum Analizi |
|-------|-------|------------|-------------|-----------|---------------|
| Operation | IT | %154 | %195 | %198 | Son 3 ayda bütçeyi neredeyse ikiye katlamış. |
| Operation | 3rd Party Service | %134 | %179 | %181 | Giderek artan bir bütçe aşımı var. |
| Admin | IT | %120 | %145 | %165 | Bütçe üstü seyretmiş. |

**ÖRNEK TABLO FORMATI (Bölüm bilgisi yoksa):**
| Kalem | Ocak Index | Şubat Index | Mart Index | Durum Analizi |
|-------|------------|-------------|-----------|---------------|
| IT | %154 | %195 | %198 | Son 3 ayda bütçeyi neredeyse ikiye katlamış. |
| 3rd Party Service | %134 | %179 | %181 | Giderek artan bir bütçe aşımı var. |

- Eğer birden fazla sayfa (sheet) varsa, soruya uygun sayfayı seç veya tüm sayfaları analiz et.
- **"SON 3 AY" İFADESİ İÇİN KRİTİK KURAL:**
  * Kullanıcı "son 3 ay" dediğinde, Excel dosyasındaki AYLARI ANALİZ ET.
  * Ayların sırasını bul (Ocak → Şubat → ... → Aralık).
  * EN SON 3 AYI kullan, ASLA ilk 3 ayı (Ocak, Şubat, Mart) kullanma.
  * Örnek: Eğer Excel'de Ağustos, Eylül, Ekim varsa → EKİM, EYLÜL, AĞUSTOS (en son 3 ay).
  * Örnek: Eğer Excel'de Ocak, Şubat, Mart, Nisan varsa → NİSAN, MART, ŞUBAT (en son 3 ay).
- Cevabı Türkçe ver.

**ÖNEMLİ: CEVAP YAPISI**
Cevabında SADECE tabloyu göster. Şunları ASLA yapma:
- "Kullanıcının sorusu:" gibi ifadeler
- "Verilen Excel dosyasının detaylı analizi..." gibi açıklamalar
- "=== BÜTÇE ANALİZİ BİLGİLERİ ===" gibi başlıklar
- "Ayların Sırası:" gibi teknik bilgiler
- "⚠️ ÖNEMLİ:" gibi uyarılar
- Soruyu tekrarlama veya özetleme

Sadece tabloyu göster. Başka hiçbir şey ekleme.

KULLANICI SORUSU: {question}

VERİ ANALİZİ RAPORU:
{excel_data}
"""

COMPARISON_PROMPT = """Sen bir Excel karşılaştırma uzmanısın. İki veri dosyasını karşılaştır ve kullanıcının sorusunu yanıtla.

KULLANICI SORUSU: {question}

GÖREVİN:
İki dosya arasındaki farkları, benzerlikleri veya istenen karşılaştırmayı analiz et.

CEVAP FORMATI:
- Sonuçları MUTLAKA bir "Markdown Tablosu" ile sun.
- Tabloda "Dosya 1 Değeri", "Dosya 2 Değeri" ve "Fark/Durum" gibi sütunlar kullan.
- Cevabı Türkçe olarak ver.

DOSYA KARŞILAŞTIRMASI:
{comparison_text}
"""

def detect_header_row(df: pd.DataFrame, max_scan_rows: int = 20) -> int:
    """
    DataFrame içindeki olası başlık satırını tespit eder.
    En çok dolu sütuna sahip olan veya belirli anahtar kelimeleri içeren satırı arar.
    """
    if df.empty:
        return 0
        
    # Yaygın başlık anahtar kelimeleri (küçük harf)
    header_keywords = [
        'tarih', 'date', 'name', 'isim', 'ad', 'soyad', 'id', 'no', 'code', 'kod',
        'amount', 'tutar', 'fiyat', 'price', 'adet', 'quantity', 'toplam', 'total',
        'bölge', 'region', 'şehir', 'city', 'ülke', 'country', 'kategori', 'category',
        'jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec',
        'ocak', 'şubat', 'mart', 'nisan', 'mayıs', 'haziran', 'temmuz', 'ağustos', 'eylül', 'ekim', 'kasım', 'aralık',
        'budget', 'bütçe', 'actual', 'hedef', 'target', 'index', 'status', 'durum', 'açıklama', 'description',
        'operation', 'subcontractor', 'renting', 'expenses'
    ]
    
    best_row_idx = 0
    max_score = -1
    
    # İlk N satırı tara
    scan_limit = min(len(df), max_scan_rows)
    
    for i in range(scan_limit):
        # Satırı string'e çevir ve küçük harfe dönüştür
        row = df.iloc[i].astype(str).str.lower()
        
        # 1. Dolu hücre sayısı skoru
        # 'nan', 'none', '', 'null' olmayan hücreleri say
        valid_cells = row.apply(lambda x: x not in ['nan', 'none', '', 'null', 'nat'] and len(x.strip()) > 0)
        non_empty_count = valid_cells.sum()
        
        # 2. Anahtar kelime eşleşme skoru
        keyword_match_count = sum(1 for val in row if any(kw in str(val) for kw in header_keywords))
        
        # Toplam skor (Anahtar kelimeler daha ağırlıklı)
        score = non_empty_count + (keyword_match_count * 3)
        
        if score > max_score:
            max_score = score
            best_row_idx = i
            
    # Eğer hiç anlamlı skor bulunamazsa ve veri azsa, 0 döndür
    if max_score <= 1 and scan_limit > 0:
        return 0
        
    return best_row_idx

def clean_and_set_header(df: pd.DataFrame) -> pd.DataFrame:
    """
    DataFrame'in başlık satırını ayarlar, temizler ve 
    MÜKERRER (DUPLICATE) SÜTUN İSİMLERİNİ BENZERSİZLEŞTİRİR.
    (Örn: 'Jan', 'Jan' -> 'Jan', 'Jan_1')
    """
    if df.empty:
        return df
        
    header_idx = detect_header_row(df)
    
    # Eğer başlık 0. satır değilse, veriyi kaydır
    if header_idx > 0:
        new_header = df.iloc[header_idx]
        df = df[header_idx + 1:].copy()
        df.columns = new_header
    
    # Sütun isimlerini string'e çevir ve temizle
    df.columns = df.columns.astype(str).str.strip()
    
    # --- KRİTİK DÜZELTME: Mükerrer (Duplicate) Sütun İsimlerini Düzeltme ---
    new_columns = []
    seen_columns = {}  # {isim: sayı}
    
    for i, col in enumerate(df.columns):
        col_name = str(col).strip()
        
        # Boş veya anlamsız isimleri düzelt
        if not col_name or col_name.lower() in ['nan', 'none', 'null', 'nat'] or col_name.lower().startswith('unnamed:'):
            # İlk birkaç satıra bakarak sütunun içeriğine göre isim öner
            sample_values = df.iloc[:5, i].dropna().astype(str).tolist()
            if sample_values:
                # Eğer sayısal değerler varsa "Değer" gibi genel bir isim
                try:
                    pd.to_numeric(sample_values)
                    col_name = f"Sütun_{i+1}"
                except:
                    # Metin değerler varsa ilk anlamlı değeri kullan veya genel isim
                    col_name = f"Sütun_{i+1}"
            else:
                col_name = f"Sütun_{i+1}"
            
        # Duplicate kontrolü (Case-insensitive değil, çünkü Jan ve jan farklı olabilir ama genellikle aynıdır)
        if col_name in seen_columns:
            seen_columns[col_name] += 1
            # Örn: "Jan" varsa ikincisi "Jan_1", üçüncüsü "Jan_2" olur
            col_name = f"{col_name}_{seen_columns[col_name]}"
        else:
            seen_columns[col_name] = 0
            
        new_columns.append(col_name)
    
    df.columns = new_columns
    
    # Boş satırları ve sütunları temizle
    df.dropna(how='all', inplace=True)
    df.dropna(how='all', axis=1, inplace=True)
    
    # Index'i sıfırla
    df.reset_index(drop=True, inplace=True)
    
    return df

def read_excel_as_dataframe(file_bytes: bytes, file_name: str = "") -> Dict[str, Any]:
    """
    Excel (.xlsx) veya CSV (.csv) dosyasını pandas DataFrame'lere çevirir.
    file_name parametresi dosya türünü anlamak için kullanılır.
    """
    dataframes = {}
    file_stream = io.BytesIO(file_bytes)
    is_csv = file_name.lower().endswith('.csv')
    
    if PANDAS_AVAILABLE:
        try:
            if is_csv:
                # CSV okuma - header=None ile oku, sonra biz bulacağız
                try:
                    df = pd.read_csv(file_stream, header=None, engine='python', encoding='utf-8-sig')
                except UnicodeDecodeError:
                    file_stream.seek(0)
                    df = pd.read_csv(file_stream, header=None, engine='python', encoding='latin-1')
                
                df = clean_and_set_header(df)
                dataframes["Sheet1"] = df
                return dataframes
            else:
                # Excel Okuma - header=None ile oku
                try:
                    # Tüm sayfaları oku
                    df_dict = pd.read_excel(file_stream, sheet_name=None, header=None)
                    
                    # Her sayfa için başlık temizliği yap
                    cleaned_dict = {}
                    for sheet, df in df_dict.items():
                        cleaned_df = clean_and_set_header(df)
                        if not cleaned_df.empty:
                            cleaned_dict[sheet] = cleaned_df
                        
                    return cleaned_dict
                except Exception as excel_error:
                    # CSV Fallback
                    print(f"⚠️ Excel okuma hatası, CSV deneniyor: {excel_error}")
                    file_stream.seek(0)
                    try:
                        df = pd.read_csv(file_stream, header=None, engine='python')
                        df = clean_and_set_header(df)
                        dataframes["Sheet1"] = df
                        return dataframes
                    except Exception as e:
                         print(f"⚠️ CSV Fallback hatası: {e}")
        except Exception as e:
            print(f"⚠️ Pandas ile okuma hatası: {e}")

    # Fallback mekanizmaları (Openpyxl ve CSV module)
    if OPENPYXL_AVAILABLE and not is_csv:
        try:
            file_stream.seek(0)
            wb = load_workbook(filename=file_stream, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    rows.append(row)
                if rows:
                    # Basit bir DataFrame yapısı taklidi - Header tespiti burada zor, ham veri dönüyoruz
                    dataframes[sheet_name] = {"rows": rows, "type": "raw"}
            return dataframes
        except Exception as e:
            print(f"❌ Openpyxl okuma hatası: {e}")
            
    try:
        file_stream.seek(0)
        try:
            content = file_stream.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            file_stream.seek(0)
            content = file_stream.read().decode('latin-1')
            
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        dataframes["Sheet1"] = {"rows": rows, "type": "raw"}
        return dataframes
    except Exception as e:
        print(f"❌ CSV module okuma hatası: {e}")

    return {}


def format_dataframe_for_llm(df: Any, sheet_name: str) -> str:
    """DataFrame'i LLM için zenginleştirilmiş metin formatına çevirir."""
    
    text = f"\n--- Sayfa: {sheet_name} ---\n"

    # Ham Veri (Dict) Durumu
    if isinstance(df, dict) and df.get("type") == "raw":
        rows = df["rows"]
        text += f"Satır sayısı: {len(rows)}\n"
        text += "\nVeri İçeriği (İlk 100 Satır):\n"
        for i, row in enumerate(rows[:100]):
            clean_row = [str(cell)[:100] if cell is not None else "" for cell in row]
            text += f"Satır {i}: | " + " | ".join(clean_row) + " |\n"
        return text
    
    # --- PANDAS DATAFRAME AKILLI ANALİZİ ---
    
    # 1. Genel Bilgiler
    text += f"Toplam Satır Sayısı: {len(df)}\n"
    columns_list = [str(col) for col in df.columns.tolist()]
    text += f"Sütunlar: {', '.join(columns_list)}\n\n"

    # Veri tiplerini ayır
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    categorical_cols = df.select_dtypes(include=['object', 'category', 'bool']).columns.tolist()

    # 2. İstatistiksel Özet
    if numeric_cols:
        try:
            description = df[numeric_cols].describe().T.to_string()
            text += "=== İSTATİSTİKSEL ÖZET (Sayısal) ===\n"
            text += f"{description}\n\n"
        except Exception as e:
            print(f"İstatistik hatası: {e}")

    # 3. Kategorik Analiz
    if categorical_cols:
        text += "=== KATEGORİK ÖZET ===\n"
        try:
            for col in categorical_cols[:5]:
                if df[col].nunique() < 20:
                    top_values = df[col].value_counts().head(5).to_string()
                    text += f"--- {col}: \n{top_values}\n"
        except Exception:
            pass

    # 4. Bütçe Analizi Bilgileri
    budget_keywords = ['budget', 'bütçe', 'annual', 'yıllık']
    month_keywords = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec',
                      'ocak', 'şubat', 'mart', 'nisan', 'mayıs', 'haziran', 'temmuz', 'ağustos', 'eylül', 'ekim', 'kasım', 'aralık']
    
    budget_cols = []
    month_cols = []
    for col in df.columns:
        col_lower = str(col).lower()
        if any(kw in col_lower for kw in budget_keywords):
            budget_cols.append(str(col))
        if any(kw in col_lower for kw in month_keywords):
            month_cols.append(str(col))
    
    if budget_cols or month_cols:
        text += "\n=== BÜTÇE ANALİZİ BİLGİLERİ ===\n"
        # Ayları sıralama mantığı...
        month_order = {
            'ocak': 1, 'january': 1, 'jan': 1, 'şubat': 2, 'february': 2, 'feb': 2,
            'mart': 3, 'march': 3, 'mar': 3, 'nisan': 4, 'april': 4, 'apr': 4,
            'mayıs': 5, 'may': 5, 'haziran': 6, 'june': 6, 'jun': 6,
            'temmuz': 7, 'july': 7, 'jul': 7, 'ağustos': 8, 'august': 8, 'aug': 8,
            'eylül': 9, 'september': 9, 'sep': 9, 'ekim': 10, 'october': 10, 'oct': 10,
            'kasım': 11, 'november': 11, 'nov': 11, 'aralık': 12, 'december': 12, 'dec': 12
        }
        
        sorted_months = []
        for col in month_cols:
            col_lower = str(col).lower()
            for month_name, month_num in month_order.items():
                if month_name in col_lower:
                    sorted_months.append((month_num, col))
                    break
        
        if sorted_months:
            sorted_months.sort(key=lambda x: x[0])
            month_names = [m[1] for m in sorted_months]
            text += f"Ayların Sırası: {' → '.join(month_names)}\n"
            if len(month_names) >= 3:
                last_3_months = month_names[-3:]
                text += f"⚠️ ÖNEMLİ: 'Son 3 ay' denirse, şu ayları kullan: {', '.join(last_3_months)}\n"

    # 5. Veri Önizleme (TABLO FORMATI - OPTİMİZE EDİLDİ)
    # Tüm veriyi göster (Operation ve Admin dahil tüm bölümler için)
    total_rows = len(df)
    preview_limit = min(1000, total_rows)  # 1000 satıra çıkarıldı (Operation ve Admin dahil)
    text += f"\n=== VERİ ÖNİZLEME (İlk {preview_limit} satır - Tüm Bölümler Dahil) ===\n"
    text += "NOT: Sayılar okunabilirlik için 2 basamağa yuvarlanmıştır.\n"
    if total_rows > preview_limit:
        text += f"NOT: Toplam {total_rows} satır var, ilk {preview_limit} satır gösteriliyor. Tüm bölümler (Operation, Admin vb.) dahil edilmeye çalışılmıştır.\n"
    
    try:
        # Satır sayısını artırdık (Operation ve Admin dahil tüm bölümleri görmek için)
        preview_df = df.head(preview_limit).copy()
        
        # Tamamen boş satırları temizle
        preview_df = preview_df.dropna(how='all')
        
        # Sütun sayısını kısıtla (24 sütun limiti)
        if len(preview_df.columns) > 24:
             cols = preview_df.columns.tolist()
             selected_cols = cols[:10] + cols[-14:]
             preview_df = preview_df[selected_cols]
             text += "NOT: Tablo çok geniş olduğu için sadece ilk 10 ve son 14 sütun gösteriliyor.\n"
        
        # --- KRİTİK GÜNCELLEME: Sayı Formatlama ---
        # Sayıları önceden yuvarlayarak LLM'e temiz veri gönderiyoruz
        for col in preview_df.columns:
            # Eğer sütun sayısal ise
            if pd.api.types.is_numeric_dtype(preview_df[col]):
                try:
                    # NaN olmayanları 2 basamaklı string'e çevir (126.7215 -> "126.72")
                    preview_df[col] = preview_df[col].apply(
                        lambda x: f"{x:.2f}" if pd.notnull(x) and isinstance(x, (int, float)) else ""
                    )
                except:
                    pass
            
            # String dönüşümü ve temizlik
            preview_df[col] = preview_df[col].astype(str)
            preview_df[col] = preview_df[col].replace(['nan', 'NaN', 'None', 'null', 'NaT', 'nat', ''], '', regex=False)
            
            # Çok uzun metinleri kırp
            preview_df[col] = preview_df[col].str.slice(0, 100)

        # Markdown tablosu oluştur
        if TABULATE_AVAILABLE:
            table_text = preview_df.to_markdown(index=True)
        else:
            table_text = preview_df.to_string(index=True)
            
        text += table_text
    except Exception as e:
        print(f"Tablo formatlama hatası: {e}")
        try:
            clean_df = df.head(100).fillna("")
            text += clean_df.to_csv(index=True)
        except:
            text += str(df.head(100).fillna(""))
    
    return text



# --- DETERMINISTIC (LLM'SİZ) BÜTÇE AŞIMI ANALİZİ ---
# Not: Bazı sorular (örn: "<Bölge> için son 3 ay bütçeyi aşan kalemler") LLM'in
# uzun veri önizlemesinde kaybolabiliyor. Bu nedenle aşağıdaki fonksiyonlar
# Excel'den doğrudan hesap yapar ve kesin tablo döndürür.

_MONTH_MAP = {
    # English
    "jan": ("Ocak", 1),
    "feb": ("Şubat", 2),
    "mar": ("Mart", 3),
    "apr": ("Nisan", 4),
    "may": ("Mayıs", 5),
    "jun": ("Haziran", 6),
    "jul": ("Temmuz", 7),
    "aug": ("Ağustos", 8),
    "sep": ("Eylül", 9),
    "oct": ("Ekim", 10),
    "nov": ("Kasım", 11),
    "dec": ("Aralık", 12),
    # Turkish (also handle ascii)
    "ocak": ("Ocak", 1),
    "subat": ("Şubat", 2),
    "şubat": ("Şubat", 2),
    "mart": ("Mart", 3),
    "nisan": ("Nisan", 4),
    "mayis": ("Mayıs", 5),
    "mayıs": ("Mayıs", 5),
    "haziran": ("Haziran", 6),
    "temmuz": ("Temmuz", 7),
    "agustos": ("Ağustos", 8),
    "ağustos": ("Ağustos", 8),
    "eylul": ("Eylül", 9),
    "eylül": ("Eylül", 9),
    "ekim": ("Ekim", 10),
    "kasim": ("Kasım", 11),
    "kasım": ("Kasım", 11),
    "aralik": ("Aralık", 12),
    "aralık": ("Aralık", 12),
}

def _normalize_month_token(s: str) -> Optional[str]:
    if not s:
        return None
    t = str(s).strip().lower()
    t = re.sub(r"\s+", "", t)
    t = t.replace(".", "")
    # "Jan_1" gibi suffix'leri at
    t = re.sub(r"[_-]\d+$", "", t)
    # "jan2024" gibi olası son ekleri temizle
    t = re.sub(r"\d+$", "", t)
    return t if t in _MONTH_MAP else None

def _detect_index_month_columns(df: pd.DataFrame) -> Dict[int, str]:
    """INDEX bloğundaki ay sütunlarını bulur ve {month_num: column_name} döndürür."""
    cols = [str(c) for c in df.columns.tolist()]
    idx_pos = None
    for i, c in enumerate(cols):
        if str(c).strip().lower() == "index" or "index" == str(c).strip().lower():
            idx_pos = i
            break
    month_cols: Dict[int, str] = {}
    # 1) INDEX kolonundan sonra ayları yakala
    if idx_pos is not None:
        for c in cols[idx_pos+1:]:
            tok = _normalize_month_token(c)
            if tok:
                month_name_tr, month_num = _MONTH_MAP[tok]
                # Aynı ay birden fazla olabilir: önce dolu olanı tercih etmek için sonradan kontrol edeceğiz
                month_cols[month_num] = str(c)
    # 2) Eğer INDEX yoksa, "index" geçen kolonları tara (fallback)
    if not month_cols:
        for c in cols:
            if "index" in str(c).lower():
                # örn: "Aug Index"
                tok = _normalize_month_token(c)
                if tok:
                    month_name_tr, month_num = _MONTH_MAP[tok]
                    month_cols[month_num] = str(c)
    return month_cols

def _pick_last_3_available_months(df: pd.DataFrame, month_cols: Dict[int, str], cat_col: Optional[str] = None) -> List[int]:
    """Veride karşılığı olan (tamamen 0/boş olmayan) ayların en son 3 tanesini seçer.

    Önemli: Bazı dosyalarda INDEX bloğunun altında kategori/kalem satırlarından sonra
    farklı özet/variance satırları gelebiliyor (ör. -1, 0.07 gibi). Bu yüzden mümkünse
    sadece kategori kolonu dolu olan satırları dikkate alır.
    """
    base_df = df
    if cat_col and cat_col in df.columns:
        tmp = df.copy()
        tmp[cat_col] = tmp[cat_col].astype(str).str.strip()
        base_df = tmp[tmp[cat_col].ne("") & tmp[cat_col].ne("nan") & tmp[cat_col].ne("None")]

    available: List[int] = []
    for m in sorted(month_cols.keys()):
        col = month_cols[m]
        if col not in df.columns:
            continue
        try:
            ser = pd.to_numeric(base_df[col], errors="coerce").fillna(0)
            # INDEX değerleri normalde 0-300 bandında olur. Negatif/çok küçük varyans satırlarını
            # ay varlığı için kriter yapmamak adına >0 olan toplamı kontrol ediyoruz.
            if float(ser[ser > 0].sum()) > 0:
                available.append(m)
        except Exception:
            if base_df[col].astype(str).str.strip().replace({"nan": "", "None": ""}).ne("").any():
                available.append(m)

    return available[-3:] if len(available) >= 3 else available

def _find_category_column(df: pd.DataFrame, sheet_name: str) -> str:
    """Kategori/Kalem kolonunu bulmaya çalışır."""
    # En iyisi sheet adı ile aynı olan kolon
    for c in df.columns:
        if str(c).strip().lower() == str(sheet_name).strip().lower():
            return str(c)
    # Yoksa ilk 'operation' olmayan object kolon
    obj_cols = df.select_dtypes(include=["object", "category"]).columns.tolist()
    for c in obj_cols:
        if str(c).strip().lower() not in ["operation", "operasyon", "op"]:
            return str(c)
    # En son fallback: ilk kolon
    return str(df.columns[0])

def _find_operation_column(df: pd.DataFrame) -> Optional[str]:
    """Bölüm/kategori kolonunu bulmaya çalışır (Operation, Admin, Factory, Production vb. gibi)."""
    # Yaygın bölüm kolonu isimleri (dosyaya özel değil, genel)
    section_keywords = [
        "operation", "operasyon", "op", 
        "bölüm", "section", "department", "dept",
        "factory", "fabrika", "production", "üretim",
        "admin", "yönetim", "management",
        "sales", "satış", "marketing", "pazarlama",
        "category", "kategori", "type", "tip"
    ]
    for c in df.columns:
        c_lower = str(c).strip().lower()
        if c_lower in section_keywords:
            return str(c)
    return None

def _safe_pct(v: Any) -> str:
    try:
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "-"
        f = float(v)
        return f"{f:.2f}"
    except Exception:
        return "-"

def _build_overbudget_markdown_table_for_last3(df, sheet_name: str) -> str:
    if df is None or df.empty:
        return ""

    month_cols = _detect_index_month_columns(df)
    if not month_cols:
        return ""

    # ✅ Kategori kolonu: bu dosyada doğru helper zaten _find_category_column
    cat_col = _find_category_column(df, sheet_name)

    last3 = _pick_last_3_available_months(df, month_cols, cat_col)
    if not last3:
        return ""

    # İlgili ay kolonları
    cols_last3 = [month_cols[m] for m in last3 if month_cols.get(m) in df.columns]
    if len(cols_last3) != 3:
        return None


    # Sayısal değerlere çevir
    work = df.copy()
    for c in cols_last3:
        work[c] = pd.to_numeric(work[c], errors="coerce")

    # Kategori boş olanları at
    work[cat_col] = work[cat_col].astype(str).str.strip()
    work = work[work[cat_col].ne("") & work[cat_col].ne("nan") & work[cat_col].ne("None")]

    # INDEX > 100: bütçe aşıldı varsayımı
    mask = (work[cols_last3[0]] > 100) | (work[cols_last3[1]] > 100) | (work[cols_last3[2]] > 100)
    over = work[mask].copy()
    if over.empty:
        return None

    # Duplicate kalemleri tekilleştir (ilk görünen)
    over = over.drop_duplicates(subset=[cat_col], keep="first")

    # Türkçe ay adları ve sütun başlıkları
    month_tr = []
    for m in last3:
        # (name_tr, num)
        # reverse lookup:
        for tok, (name_tr, num) in _MONTH_MAP.items():
            if num == m:
                month_tr.append(name_tr)
                break
        else:
            month_tr.append(str(m))

    # Analiz cümlesi üret
    def make_comment(row) -> str:
        vals = [row[cols_last3[0]], row[cols_last3[1]], row[cols_last3[2]]]
        flags = [v is not None and not (isinstance(v, float) and np.isnan(v)) and v > 100 for v in vals]
        # Üç ay da aşıyorsa
        if all(flags):
            return "Son 3 ayda bütçeyi sürekli aşmış."
        # Sadece son ay (en güncel) aşıyorsa
        if flags[2] and not flags[0] and not flags[1]:
            return f"{month_tr[2]} ayında bütçeyi önemli ölçüde aşmış."
        # Artan trend
        try:
            if all(v is not None and not (isinstance(v, float) and np.isnan(v)) for v in vals):
                if vals[0] < vals[1] < vals[2] and vals[2] > 100:
                    return f"Giderek artan bir bütçe aşımı var, {month_tr[2]} ayında zirve yapmış."
        except Exception:
            pass
        # Genel
        return "Son 3 ayda bütçe üstü seyrettiği ay(lar) var."

    # Bölüm kolonunu bul (Operation, Admin, Factory, Production vb. gibi)
    operation_col = _find_operation_column(df)
    
    out_rows = []
    for _, r in over.iterrows():
        row_data = {}
        # Eğer bölüm kolonu varsa, Bölüm bilgisini ilk sütun olarak ekle
        if operation_col and operation_col in r.index:
            operation_value = str(r[operation_col]).strip()
            if operation_value and operation_value.lower() not in ["nan", "none", ""]:
                row_data["Bölüm"] = operation_value
        
        # Diğer sütunlar
        row_data["Kalem"] = r[cat_col]
        row_data[f"{month_tr[0]} Index"] = _safe_pct(r[cols_last3[0]])
        row_data[f"{month_tr[1]} Index"] = _safe_pct(r[cols_last3[1]])
        row_data[f"{month_tr[2]} Index"] = _safe_pct(r[cols_last3[2]])
        row_data["Durum Analizi"] = make_comment(r)
        
        out_rows.append(row_data)

    out_df = pd.DataFrame(out_rows)

    # Markdown üret
    if TABULATE_AVAILABLE:
        md = tabulate.tabulate(out_df, headers="keys", tablefmt="github", showindex=False)
    else:
        md = out_df.to_markdown(index=False)

    return md


def analyze_excel_data(
    file_bytes: bytes,
    question: str,
    model_name: str = "gemini",
    file_name: str = "data.xlsx" 
) -> str:
    """Excel/CSV dosyasını analiz eder ve soruyu yanıtlar."""
    try:
        dataframes = read_excel_as_dataframe(file_bytes, file_name)
        if _is_overbudget_question(question):
            # Hedef sheet seçimi: Soru içinde geçen sheet adını bul, yoksa ilk sheet'i kullan
            target_sheet = None
            q_lower = (question or "").lower()
            for s in dataframes.keys():
                s_lower = str(s).strip().lower()
                if s_lower in q_lower:
                    target_sheet = s
                    break
            if target_sheet is None:
                # yoksa ilk sheet
                target_sheet = list(dataframes.keys())[0]

            df = dataframes[target_sheet]
            # Dosyadaki en güncel 3 ayı tespit et (senin mevcut helper’larınla)
            month_cols = _detect_index_month_columns(df)
            cat_col = _find_category_column(df, str(target_sheet))
            last3_nums = _pick_last_3_available_months(df, month_cols, cat_col)

            # num -> isim
            last3_names = []
            for m in last3_nums:
                for tok, (name_tr, num) in _MONTH_MAP.items():
                    if num == m:
                        last3_names.append(name_tr)
                        break

            needs_month = ("son 3 ay" in (question or "").lower()) and (not _question_mentions_months(question))
            needs_rule = (not _question_defines_overbudget_rule(question))

            if needs_month or needs_rule:
                return _build_clarification_message(str(target_sheet), last3_names)

        # --- ÖNCE DETERMINISTIC ÇÖZÜMÜ DENE (LLM'SİZ) ---
        # Özellikle: "<Bölge> için son 3 ayda bütçeyi aşan/uygun olmayan kalemler" soruları.
        q_lower = (question or "").lower()
        wants_last3 = "son 3 ay" in q_lower or "last 3 month" in q_lower or "last three month" in q_lower
        wants_budget_over = ("bütçe" in q_lower or "budget" in q_lower) and (
            "aş" in q_lower or "over" in q_lower or "uygun olmayan" in q_lower or "exceed" in q_lower
        )
        if wants_last3 and wants_budget_over:
            # Hedef sheet seçimi: soru içinde geçen sheet adı varsa onu kullan, yoksa tek sheet ise onu kullan
            target_sheet = None
            for sname in dataframes.keys():
                if str(sname).lower() in q_lower:
                    target_sheet = sname
                    break
            if target_sheet is None and len(dataframes) == 1:
                target_sheet = list(dataframes.keys())[0]
            # Eğer soru bölge içeriyorsa ama sheet yoksa, dosyadaki sheet isimlerini kullanarak eşleştir
            if target_sheet is None:
                # Dosyadaki tüm sheet isimlerini al
                available_sheets = [str(sname).lower() for sname in dataframes.keys()]
                # Soru içinde geçen kelimeleri kontrol et
                question_words = q_lower.split()
                for word in question_words:
                    # Her sheet ismiyle karşılaştır
                    for sname in dataframes.keys():
                        sname_lower = str(sname).lower()
                        # Tam eşleşme veya kısmi eşleşme kontrolü
                        if word in sname_lower or sname_lower in word:
                            target_sheet = sname
                            break
                    if target_sheet:
                        break

            if target_sheet and not (isinstance(dataframes[target_sheet], dict) and dataframes[target_sheet].get("type") == "raw"):
                md = _build_overbudget_markdown_table_for_last3(dataframes[target_sheet], str(target_sheet))
                if md:
                    return md
        
        if not dataframes:
            return "Dosya okunamadı. Lütfen geçerli bir Excel veya CSV dosyası yükleyin."
        
        # Birden fazla sheet varsa bilgi ver
        sheet_count = len(dataframes)
        excel_summary = f"=== EXCEL DOSYASI BİLGİLERİ ===\n"
        excel_summary += f"Dosya Adı: {file_name}\n"
        excel_summary += f"Toplam Sayfa Sayısı: {sheet_count}\n"
        excel_summary += f"Sayfa İsimleri: {', '.join(dataframes.keys())}\n\n"
        
        if sheet_count > 1:
            excel_summary += "NOT: Bu Excel dosyasında birden fazla sayfa (sheet) bulunmaktadır. "
            excel_summary += "Soruya uygun sayfayı seçerek analiz yapmalısın. "
            excel_summary += "Eğer soru tüm sayfaları kapsıyorsa, tüm sayfaları analiz et.\n\n"
        
        excel_summary += "=== SAYFA İÇERİKLERİ ===\n\n"
        
        for sheet_name, df in dataframes.items():
            excel_summary += format_dataframe_for_llm(df, sheet_name)
            excel_summary += "\n"  # Sayfalar arası boşluk
        
        # Token sınırını aşmamak için özetin çok uzun olmadığından emin ol
        if len(excel_summary) > 80000:
            excel_summary = excel_summary[:80000] + "\n...(Veri çok uzun olduğu için kesildi)..."

        llm = get_llm_for_model(model_name)
        prompt = EXCEL_AGENT_PROMPT.format(
            question=question,
            excel_data=excel_summary
        )
        
        response = llm.invoke(prompt)
        return response.content if hasattr(response, 'content') else str(response)
        
    except Exception as e:
        print(f"❌ Analiz hatası: {e}")
        traceback.print_exc()
        return f"Analiz sırasında hata oluştu: {str(e)}"


def compare_excel_files(
    file1_bytes: bytes,
    file2_bytes: bytes,
    question: str,
    model_name: str = "gemini",
    file1_name: str = "file1.xlsx", 
    file2_name: str = "file2.xlsx" 
) -> str:
    """İki Excel dosyasını karşılaştırır."""
    try:
        df1_dict = read_excel_as_dataframe(file1_bytes, file1_name)
        df2_dict = read_excel_as_dataframe(file2_bytes, file2_name)
        
        if not df1_dict:
            return f"İlk dosya ({file1_name}) okunamadı."
        if not df2_dict:
            return f"İkinci dosya ({file2_name}) okunamadı."
        
        comparison_text = f"=== İLK DOSYA: {file1_name} ===\n"
        for sheet_name, df in df1_dict.items():
            comparison_text += format_dataframe_for_llm(df, sheet_name)
        
        comparison_text += f"\n=== İKİNCİ DOSYA: {file2_name} ===\n"
        for sheet_name, df in df2_dict.items():
            comparison_text += format_dataframe_for_llm(df, sheet_name)
        
        llm = get_llm_for_model(model_name)
        prompt = COMPARISON_PROMPT.format(
            question=question,
            comparison_text=comparison_text
        )
        
        response = llm.invoke(prompt)
        return response.content if hasattr(response, 'content') else str(response)
        
    except Exception as e:
        print(f"❌ Excel karşılaştırma hatası: {e}")
        traceback.print_exc()
        return f"Excel karşılaştırma sırasında hata oluştu: {str(e)}"
def _is_overbudget_question(q: str) -> bool:
    q = (q or "").lower()
    keys = ["son 3 ay", "last 3", "bütçe", "budget", "uygun değil", "aş", "over", "exceed", "index"]
    return sum(k in q for k in keys) >= 2

def _question_mentions_months(q: str) -> bool:
    q = (q or "").lower()
    month_tokens = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec",
                    "ocak","şubat","mart","nisan","mayıs","haziran","temmuz","ağustos","eylül","ekim","kasım","aralık"]
    return any(m in q for m in month_tokens)

def _question_defines_overbudget_rule(q: str) -> bool:
    q = (q or "").lower()
    # Kullanıcı açıkça index>100 ya da actual/budget dedi mi?
    return ("index" in q) or ("> 100" in q) or ("actual" in q) or ("budget/12" in q) or ("bütçe/12" in q)

def _build_clarification_message(sheet_name: str, last3_month_names: list[str]) -> str:
    months = "–".join(last3_month_names) if last3_month_names else "?"
    return (
        f"Dosyayı açtım, **{sheet_name}** sekmesini inceledim 👍\n"
        "Ancak net ve doğru liste çıkarabilmem için küçük ama kritik bir netleştirme gerekiyor.\n\n"
        "**1) Son 3 ay hangileri?**\n"
        f"- Önerim (dosyadaki en güncel 3 ay): **{months}**\n\n"
        "**2) “Bütçeye uygun değil” ne demek?**\n"
        "a) Aylık **Actual > Annual Budget/12**\n"
        "b) **Index 100 > 100** olan kalemler\n"
        "c) **Her ikisi de** (hangisi varsa)\n\n"
        "Kısa cevap yazman yeterli: örn. `1:{months} 2:b`"
    )
