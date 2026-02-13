# backend/app/services/excel_agent_service.py
# GÜNCELLENMİŞ VERSİYON: Dinamik Header Tespiti + Duplicate Sütun Düzeltme + Akıllı Analiz

import io
import csv
import traceback
from typing import Dict, Any, Optional, List
from app.services.llm_providers import get_llm_for_model

# Pandas ve openpyxl import'ları
try:
    import pandas as pd
    import numpy as np
    PANDAS_AVAILABLE = True
except ImportError:
    print("UYARI: 'pandas' kütüphanesi yüklü değil. Excel analizi sınırlı olacak.")
    PANDAS_AVAILABLE = False
    pd = None

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("UYARI: 'openpyxl' kütüphanesi yüklü değil. Excel okuma çalışmayacak.")
    OPENPYXL_AVAILABLE = False
    load_workbook = None

# Tabulate kontrolü (Markdown tabloları için)
try:
    import tabulate
    TABULATE_AVAILABLE = True
except ImportError:
    TABULATE_AVAILABLE = False

# --- GÜNCELLENMİŞ PROMPTLAR ---

# --- GÜNCELLENMİŞ PROMPT (LİMİTSİZ) ---

EXCEL_AGENT_PROMPT = """Sen uzman bir Veri Analistisin. Sana bir Excel/CSV dosyasının detaylı analizi, istatistikleri ve veri önizlemesi verildi.

GÖREVİN:
Kullanıcının sorusunu, AŞAĞIDAKİ VERİLERİ kullanarak yanıtlamak.

ÖNEMLİ TARİH KURALLARI:
- "Son 3 ay" ifadesi kullanıldığında, Excel dosyasındaki AYLARI ANALİZ ET ve EN SON 3 AYI bul.
- Aylar genellikle şu sırada olur: Ocak, Şubat, Mart, Nisan, Mayıs, Haziran, Temmuz, Ağustos, Eylül, Ekim, Kasım, Aralık
- İngilizce aylar: January, February, March, April, May, June, July, August, September, October, November, December
- Kısa formlar: Jan, Feb, Mar, Apr, May, Jun, Jul, Aug, Sep, Oct, Nov, Dec
- ÖRNEK: Eğer Excel'de Ağustos, Eylül, Ekim, Kasım, Aralık varsa ve "son 3 ay" denirse, EKİM, KASIM, ARALIK'ı kullan (en son 3 ay).
- ÖRNEK: Eğer Excel'de Ocak, Şubat, Mart, Nisan varsa ve "son 3 ay" denirse, ŞUBAT, MART, NİSAN'ı kullan (en son 3 ay).
- ASLA ilk 3 ayı (Ocak, Şubat, Mart) kullanma, EN SON 3 AYI kullan.

KULLANILACAK KAYNAKLAR:
1. **"OTOMATİK GRUPLAMA ANALİZİ" (EN ÖNEMLİ):** Hazır hesaplanmış özetler buradadır.
2. **"VERİ ÖNİZLEME":** Tabloyu satır satır incelemek için burayı kullan.

CEVAP FORMATI KURALLARI (ÇOK ÖNEMLİ):
- Cevaplarını sunarken MUTLAKA **Markdown Tablosu** kullan.
- **ASLA** sana verilen "VERİ ÖNİZLEME" tablosunun tamamını kopyalayıp yapıştırma.
- Sadece sorulan soruya (örneğin "Bütçeye uygun olmayanlar") uyan satırları filtreleyerek yeni bir tablo oluştur.
- Tabloda gereksiz tekrar eden satırlar oluşturma. Her kalemden sadece bir tane olsun.
- **LİMİT YOKTUR:** Şartları sağlayan 10, 20 veya 50 kalem varsa **HEPSİNİ TABLOYA EKLE.**
- **ASLA** "ve diğerleri..." diyerek listeyi yarıda kesme. Tabloyu tam ve eksiksiz ver.

**TABLO FORMATI KURALLARI (KRİTİK - MUTLAKA UYGULA):**
- Tabloda "nan", "NaN", "None", "null" gibi teknik değerler ASLA gösterilmemeli. Boş hücreler için "-" kullan.
- Sütun başlıkları anlamlı ve açıklayıcı olmalı. "Column_2", "nan" gibi teknik isimler kullanma.
- **BÜTÇE ANALİZİ TABLOSU FORMATI:**
  * İlk sütun: "Kalem" veya "Kategori" - Harcama kalemlerinin isimleri (örn: "IT", "3rd Party Service", "Repair & Maintenance")
  * Orta sütunlar: Aylık değerler veya index'ler (örn: "Ocak", "Şubat", "Mart" veya "Ocak Index", "Şubat Index")
    - Eğer index/yüzde ise "%" işareti kullan (örn: %154, %195)
    - Eğer tutar ise sayısal değer göster (örn: 1,234.50)
  * Son sütun: "Durum Analizi" veya "Açıklama" - Her kalem için kısa bir analiz metni
    - Örnek: "Son 3 ayda bütçeyi neredeyse ikiye katlamış."
    - Örnek: "Giderek artan bir bütçe aşımı var."
    - Örnek: "Sürekli bütçe üstünde, Ekim ayında artış göstermiş."
- Sayısal değerler düzgün formatlanmalı:
  * Yüzdeler için: %154, %195.5 gibi
  * Tutarlar için: 1,234.50 veya 1234.50 gibi
  * Ondalık sayılar için maksimum 2 basamak göster
- Tablo düzeni temiz ve okunabilir olmalı:
  * Sütunlar hizalı
  * Başlıklar net ve anlaşılır
  * Her satır bir kalemi temsil etmeli
  * Tekrarlayan satırlar olmamalı

**ÖRNEK TABLO FORMATI:**
| Kalem | Ocak Index | Şubat Index | Mart Index | Durum Analizi |
|-------|------------|-------------|-----------|---------------|
| IT | %154 | %195 | %198 | Son 3 ayda bütçeyi neredeyse ikiye katlamış. |
| 3rd Party Service | %134 | %179 | %181 | Giderek artan bir bütçe aşımı var. |

- Eğer birden fazla sayfa (sheet) varsa, soruya uygun sayfayı seç veya tüm sayfaları analiz et.
- **"SON 3 AY" İFADESİ İÇİN KRİTİK KURAL:**
  * Kullanıcı "son 3 ay" dediğinde, Excel dosyasındaki AYLARI ANALİZ ET.
  * Ayların sırasını bul (Ocak → Şubat → ... → Aralık).
  * EN SON 3 AYI kullan, ASLA ilk 3 ayı (Ocak, Şubat, Mart) kullanma.
  * Örnek: Eğer Excel'de Ağustos, Eylül, Ekim varsa → EKİM, EYLÜL, AĞUSTOS (en son 3 ay).
  * Örnek: Eğer Excel'de Ocak, Şubat, Mart, Nisan varsa → NİSAN, MART, ŞUBAT (en son 3 ay).
- Cevabı Türkçe ver.

KULLANICI SORUSU: {question}

VERİ ANALİZİ RAPORU:
{excel_data}
"""

COMPARISON_PROMPT = """Sen bir Excel karşılaştırma uzmanısın. İki veri dosyasını karşılaştır ve kullanıcının sorusunu yanıtla.

KULLANICI SORUSU: {question}

GÖREVİN:
İki dosya arasındaki farkları, benzerlikleri veya istenen karşılaştırmayı analiz et.

CEVAP FORMATI:
- Sonuçları MUTLAKA bir "Markdown Tablosu" ile sun.
- Tabloda "Dosya 1 Değeri", "Dosya 2 Değeri" ve "Fark/Durum" gibi sütunlar kullan.
- Cevabı Türkçe olarak ver.

DOSYA KARŞILAŞTIRMASI:
{comparison_text}
"""

def detect_header_row(df: pd.DataFrame, max_scan_rows: int = 20) -> int:
    """
    DataFrame içindeki olası başlık satırını tespit eder.
    En çok dolu sütuna sahip olan veya belirli anahtar kelimeleri içeren satırı arar.
    """
    if df.empty:
        return 0
        
    # Yaygın başlık anahtar kelimeleri (küçük harf)
    header_keywords = [
        'tarih', 'date', 'name', 'isim', 'ad', 'soyad', 'id', 'no', 'code', 'kod',
        'amount', 'tutar', 'fiyat', 'price', 'adet', 'quantity', 'toplam', 'total',
        'bölge', 'region', 'şehir', 'city', 'ülke', 'country', 'kategori', 'category',
        'jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec',
        'ocak', 'şubat', 'mart', 'nisan', 'mayıs', 'haziran', 'temmuz', 'ağustos', 'eylül', 'ekim', 'kasım', 'aralık',
        'budget', 'bütçe', 'actual', 'hedef', 'target', 'index', 'status', 'durum', 'açıklama', 'description',
        'operation', 'subcontractor', 'renting', 'expenses'
    ]
    
    best_row_idx = 0
    max_score = -1
    
    # İlk N satırı tara
    scan_limit = min(len(df), max_scan_rows)
    
    for i in range(scan_limit):
        # Satırı string'e çevir ve küçük harfe dönüştür
        row = df.iloc[i].astype(str).str.lower()
        
        # 1. Dolu hücre sayısı skoru
        # 'nan', 'none', '', 'null' olmayan hücreleri say
        valid_cells = row.apply(lambda x: x not in ['nan', 'none', '', 'null', 'nat'] and len(x.strip()) > 0)
        non_empty_count = valid_cells.sum()
        
        # 2. Anahtar kelime eşleşme skoru
        keyword_match_count = sum(1 for val in row if any(kw in str(val) for kw in header_keywords))
        
        # Toplam skor (Anahtar kelimeler daha ağırlıklı)
        score = non_empty_count + (keyword_match_count * 3)
        
        if score > max_score:
            max_score = score
            best_row_idx = i
            
    # Eğer hiç anlamlı skor bulunamazsa ve veri azsa, 0 döndür
    if max_score <= 1 and scan_limit > 0:
        return 0
        
    return best_row_idx

def clean_and_set_header(df: pd.DataFrame) -> pd.DataFrame:
    """
    DataFrame'in başlık satırını ayarlar, temizler ve 
    MÜKERRER (DUPLICATE) SÜTUN İSİMLERİNİ BENZERSİZLEŞTİRİR.
    (Örn: 'Jan', 'Jan' -> 'Jan', 'Jan_1')
    """
    if df.empty:
        return df
        
    header_idx = detect_header_row(df)
    
    # Eğer başlık 0. satır değilse, veriyi kaydır
    if header_idx > 0:
        new_header = df.iloc[header_idx]
        df = df[header_idx + 1:].copy()
        df.columns = new_header
    
    # Sütun isimlerini string'e çevir ve temizle
    df.columns = df.columns.astype(str).str.strip()
    
    # --- KRİTİK DÜZELTME: Mükerrer (Duplicate) Sütun İsimlerini Düzeltme ---
    new_columns = []
    seen_columns = {}  # {isim: sayı}
    
    for i, col in enumerate(df.columns):
        col_name = str(col).strip()
        
        # Boş veya anlamsız isimleri düzelt
        if not col_name or col_name.lower() in ['nan', 'none', 'null', 'nat'] or col_name.lower().startswith('unnamed:'):
            # İlk birkaç satıra bakarak sütunun içeriğine göre isim öner
            sample_values = df.iloc[:5, i].dropna().astype(str).tolist()
            if sample_values:
                # Eğer sayısal değerler varsa "Değer" gibi genel bir isim
                try:
                    pd.to_numeric(sample_values)
                    col_name = f"Sütun_{i+1}"
                except:
                    # Metin değerler varsa ilk anlamlı değeri kullan veya genel isim
                    col_name = f"Sütun_{i+1}"
            else:
                col_name = f"Sütun_{i+1}"
            
        # Duplicate kontrolü (Case-insensitive değil, çünkü Jan ve jan farklı olabilir ama genellikle aynıdır)
        if col_name in seen_columns:
            seen_columns[col_name] += 1
            # Örn: "Jan" varsa ikincisi "Jan_1", üçüncüsü "Jan_2" olur
            col_name = f"{col_name}_{seen_columns[col_name]}"
        else:
            seen_columns[col_name] = 0
            
        new_columns.append(col_name)
    
    df.columns = new_columns
    
    # Boş satırları ve sütunları temizle
    df.dropna(how='all', inplace=True)
    df.dropna(how='all', axis=1, inplace=True)
    
    # Index'i sıfırla
    df.reset_index(drop=True, inplace=True)
    
    return df

def read_excel_as_dataframe(file_bytes: bytes, file_name: str = "") -> Dict[str, Any]:
    """
    Excel (.xlsx) veya CSV (.csv) dosyasını pandas DataFrame'lere çevirir.
    file_name parametresi dosya türünü anlamak için kullanılır.
    """
    dataframes = {}
    file_stream = io.BytesIO(file_bytes)
    is_csv = file_name.lower().endswith('.csv')
    
    if PANDAS_AVAILABLE:
        try:
            if is_csv:
                # CSV okuma - header=None ile oku, sonra biz bulacağız
                try:
                    df = pd.read_csv(file_stream, header=None, engine='python', encoding='utf-8-sig')
                except UnicodeDecodeError:
                    file_stream.seek(0)
                    df = pd.read_csv(file_stream, header=None, engine='python', encoding='latin-1')
                
                df = clean_and_set_header(df)
                dataframes["Sheet1"] = df
                return dataframes
            else:
                # Excel Okuma - header=None ile oku
                try:
                    # Tüm sayfaları oku
                    df_dict = pd.read_excel(file_stream, sheet_name=None, header=None)
                    
                    # Her sayfa için başlık temizliği yap
                    cleaned_dict = {}
                    for sheet, df in df_dict.items():
                        cleaned_df = clean_and_set_header(df)
                        if not cleaned_df.empty:
                            cleaned_dict[sheet] = cleaned_df
                        
                    return cleaned_dict
                except Exception as excel_error:
                    # CSV Fallback
                    print(f"⚠️ Excel okuma hatası, CSV deneniyor: {excel_error}")
                    file_stream.seek(0)
                    try:
                        df = pd.read_csv(file_stream, header=None, engine='python')
                        df = clean_and_set_header(df)
                        dataframes["Sheet1"] = df
                        return dataframes
                    except Exception as e:
                         print(f"⚠️ CSV Fallback hatası: {e}")
        except Exception as e:
            print(f"⚠️ Pandas ile okuma hatası: {e}")

    # Fallback mekanizmaları (Openpyxl ve CSV module)
    if OPENPYXL_AVAILABLE and not is_csv:
        try:
            file_stream.seek(0)
            wb = load_workbook(filename=file_stream, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    rows.append(row)
                if rows:
                    # Basit bir DataFrame yapısı taklidi - Header tespiti burada zor, ham veri dönüyoruz
                    dataframes[sheet_name] = {"rows": rows, "type": "raw"}
            return dataframes
        except Exception as e:
            print(f"❌ Openpyxl okuma hatası: {e}")
            
    try:
        file_stream.seek(0)
        try:
            content = file_stream.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            file_stream.seek(0)
            content = file_stream.read().decode('latin-1')
            
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        dataframes["Sheet1"] = {"rows": rows, "type": "raw"}
        return dataframes
    except Exception as e:
        print(f"❌ CSV module okuma hatası: {e}")

    return {}


def format_dataframe_for_llm(df: Any, sheet_name: str) -> str:
    """DataFrame'i LLM için zenginleştirilmiş metin formatına çevirir."""
    
    text = f"\n--- Sayfa: {sheet_name} ---\n"

    # Ham Veri (Dict) Durumu
    if isinstance(df, dict) and df.get("type") == "raw":
        rows = df["rows"]
        text += f"Satır sayısı: {len(rows)}\n"
        text += "\nVeri İçeriği (İlk 100 Satır):\n"
        for i, row in enumerate(rows[:100]):
            clean_row = [str(cell)[:100] if cell is not None else "" for cell in row]
            text += f"Satır {i}: | " + " | ".join(clean_row) + " |\n"
        return text
    
    # --- PANDAS DATAFRAME AKILLI ANALİZİ ---
    
    # 1. Genel Bilgiler
    text += f"Toplam Satır Sayısı: {len(df)}\n"
    columns_list = [str(col) for col in df.columns.tolist()]
    text += f"Sütunlar: {', '.join(columns_list)}\n\n"

    # Veri tiplerini ayır
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    categorical_cols = df.select_dtypes(include=['object', 'category', 'bool']).columns.tolist()

    # 2. İstatistiksel Özet
    if numeric_cols:
        try:
            description = df[numeric_cols].describe().T.to_string()
            text += "=== İSTATİSTİKSEL ÖZET (Sayısal) ===\n"
            text += f"{description}\n\n"
        except Exception as e:
            print(f"İstatistik hatası: {e}")

    # 3. Kategorik Analiz
    if categorical_cols:
        text += "=== KATEGORİK ÖZET ===\n"
        try:
            for col in categorical_cols[:5]:
                if df[col].nunique() < 20:
                    top_values = df[col].value_counts().head(5).to_string()
                    text += f"--- {col}: \n{top_values}\n"
        except Exception:
            pass

    # 4. Bütçe Analizi Bilgileri
    budget_keywords = ['budget', 'bütçe', 'annual', 'yıllık']
    month_keywords = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec',
                      'ocak', 'şubat', 'mart', 'nisan', 'mayıs', 'haziran', 'temmuz', 'ağustos', 'eylül', 'ekim', 'kasım', 'aralık']
    
    budget_cols = []
    month_cols = []
    for col in df.columns:
        col_lower = str(col).lower()
        if any(kw in col_lower for kw in budget_keywords):
            budget_cols.append(str(col))
        if any(kw in col_lower for kw in month_keywords):
            month_cols.append(str(col))
    
    if budget_cols or month_cols:
        text += "\n=== BÜTÇE ANALİZİ BİLGİLERİ ===\n"
        # Ayları sıralama mantığı...
        month_order = {
            'ocak': 1, 'january': 1, 'jan': 1, 'şubat': 2, 'february': 2, 'feb': 2,
            'mart': 3, 'march': 3, 'mar': 3, 'nisan': 4, 'april': 4, 'apr': 4,
            'mayıs': 5, 'may': 5, 'haziran': 6, 'june': 6, 'jun': 6,
            'temmuz': 7, 'july': 7, 'jul': 7, 'ağustos': 8, 'august': 8, 'aug': 8,
            'eylül': 9, 'september': 9, 'sep': 9, 'ekim': 10, 'october': 10, 'oct': 10,
            'kasım': 11, 'november': 11, 'nov': 11, 'aralık': 12, 'december': 12, 'dec': 12
        }
        
        sorted_months = []
        for col in month_cols:
            col_lower = str(col).lower()
            for month_name, month_num in month_order.items():
                if month_name in col_lower:
                    sorted_months.append((month_num, col))
                    break
        
        if sorted_months:
            sorted_months.sort(key=lambda x: x[0])
            month_names = [m[1] for m in sorted_months]
            text += f"Ayların Sırası: {' → '.join(month_names)}\n"
            if len(month_names) >= 3:
                last_3_months = month_names[-3:]
                text += f"⚠️ ÖNEMLİ: 'Son 3 ay' denirse, şu ayları kullan: {', '.join(last_3_months)}\n"

    # 5. Veri Önizleme (TABLO FORMATI - OPTİMİZE EDİLDİ)
    text += "\n=== VERİ ÖNİZLEME (İlk 300 satır - Optimize Edilmiş) ===\n"
    text += "NOT: Sayılar okunabilirlik için 2 basamağa yuvarlanmıştır.\n"
    
    try:
        # Satır sayısını 300'de tutuyoruz (Admin'i görmek için)
        preview_df = df.head(1000).copy()
        
        # Tamamen boş satırları temizle
        preview_df = preview_df.dropna(how='all')
        
        # Sütun sayısını kısıtla (24 sütun limiti)
        if len(preview_df.columns) > 24:
             cols = preview_df.columns.tolist()
             selected_cols = cols[:10] + cols[-14:]
             preview_df = preview_df[selected_cols]
             text += "NOT: Tablo çok geniş olduğu için sadece ilk 10 ve son 14 sütun gösteriliyor.\n"
        
        # --- KRİTİK GÜNCELLEME: Sayı Formatlama ---
        # Sayıları önceden yuvarlayarak LLM'e temiz veri gönderiyoruz
        for col in preview_df.columns:
            # Eğer sütun sayısal ise
            if pd.api.types.is_numeric_dtype(preview_df[col]):
                try:
                    # NaN olmayanları 2 basamaklı string'e çevir (126.7215 -> "126.72")
                    preview_df[col] = preview_df[col].apply(
                        lambda x: f"{x:.2f}" if pd.notnull(x) and isinstance(x, (int, float)) else ""
                    )
                except:
                    pass
            
            # String dönüşümü ve temizlik
            preview_df[col] = preview_df[col].astype(str)
            preview_df[col] = preview_df[col].replace(['nan', 'NaN', 'None', 'null', 'NaT', 'nat', ''], '', regex=False)
            
            # Çok uzun metinleri kırp
            preview_df[col] = preview_df[col].str.slice(0, 100)

        # Markdown tablosu oluştur
        if TABULATE_AVAILABLE:
            table_text = preview_df.to_markdown(index=True)
        else:
            table_text = preview_df.to_string(index=True)
            
        text += table_text
    except Exception as e:
        print(f"Tablo formatlama hatası: {e}")
        try:
            clean_df = df.head(100).fillna("")
            text += clean_df.to_csv(index=True)
        except:
            text += str(df.head(100).fillna(""))
    
    return text


def analyze_excel_data(
    file_bytes: bytes,
    question: str,
    model_name: str = "gemini",
    file_name: str = "data.xlsx" 
) -> str:
    """Excel/CSV dosyasını analiz eder ve soruyu yanıtlar."""
    try:
        dataframes = read_excel_as_dataframe(file_bytes, file_name)
        
        if not dataframes:
            return "Dosya okunamadı. Lütfen geçerli bir Excel veya CSV dosyası yükleyin."
        
        # Birden fazla sheet varsa bilgi ver
        sheet_count = len(dataframes)
        excel_summary = f"=== EXCEL DOSYASI BİLGİLERİ ===\n"
        excel_summary += f"Dosya Adı: {file_name}\n"
        excel_summary += f"Toplam Sayfa Sayısı: {sheet_count}\n"
        excel_summary += f"Sayfa İsimleri: {', '.join(dataframes.keys())}\n\n"
        
        if sheet_count > 1:
            excel_summary += "NOT: Bu Excel dosyasında birden fazla sayfa (sheet) bulunmaktadır. "
            excel_summary += "Soruya uygun sayfayı seçerek analiz yapmalısın. "
            excel_summary += "Eğer soru tüm sayfaları kapsıyorsa, tüm sayfaları analiz et.\n\n"
        
        excel_summary += "=== SAYFA İÇERİKLERİ ===\n\n"
        
        for sheet_name, df in dataframes.items():
            excel_summary += format_dataframe_for_llm(df, sheet_name)
            excel_summary += "\n"  # Sayfalar arası boşluk
        
        # Token sınırını aşmamak için özetin çok uzun olmadığından emin ol
        if len(excel_summary) > 80000:
            excel_summary = excel_summary[:80000] + "\n...(Veri çok uzun olduğu için kesildi)..."

        llm = get_llm_for_model(model_name)
        prompt = EXCEL_AGENT_PROMPT.format(
            question=question,
            excel_data=excel_summary
        )
        
        response = llm.invoke(prompt)
        return response.content if hasattr(response, 'content') else str(response)
        
    except Exception as e:
        print(f"❌ Analiz hatası: {e}")
        traceback.print_exc()
        return f"Analiz sırasında hata oluştu: {str(e)}"


def compare_excel_files(
    file1_bytes: bytes,
    file2_bytes: bytes,
    question: str,
    model_name: str = "gemini",
    file1_name: str = "file1.xlsx", 
    file2_name: str = "file2.xlsx" 
) -> str:
    """İki Excel dosyasını karşılaştırır."""
    try:
        df1_dict = read_excel_as_dataframe(file1_bytes, file1_name)
        df2_dict = read_excel_as_dataframe(file2_bytes, file2_name)
        
        if not df1_dict:
            return f"İlk dosya ({file1_name}) okunamadı."
        if not df2_dict:
            return f"İkinci dosya ({file2_name}) okunamadı."
        
        comparison_text = f"=== İLK DOSYA: {file1_name} ===\n"
        for sheet_name, df in df1_dict.items():
            comparison_text += format_dataframe_for_llm(df, sheet_name)
        
        comparison_text += f"\n=== İKİNCİ DOSYA: {file2_name} ===\n"
        for sheet_name, df in df2_dict.items():
            comparison_text += format_dataframe_for_llm(df, sheet_name)
        
        llm = get_llm_for_model(model_name)
        prompt = COMPARISON_PROMPT.format(
            question=question,
            comparison_text=comparison_text
        )
        
        response = llm.invoke(prompt)
        return response.content if hasattr(response, 'content') else str(response)
        
    except Exception as e:
        print(f"❌ Excel karşılaştırma hatası: {e}")
        traceback.print_exc()
        return f"Excel karşılaştırma sırasında hata oluştu: {str(e)}"